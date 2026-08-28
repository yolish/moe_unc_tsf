"""Render the ETT expert sweep (section 6 of docs/calibration_results_tsf.md) as csv + xlsx.

Reads docs/calibration_results_tsf.csv -- so run scripts/collect_calibration_results.py first --
and writes, with the same scope and the same aggregation as the markdown section:

    docs/ett_expert_sweep_mog.csv    tidy long format, one row per (pred_len, dataset, method,
                                     num_experts); dataset='ALL' is the 4-dataset mean
    docs/ett_expert_sweep_mog.xlsx   the same numbers laid out as wide tables, one sheet per
                                     horizon, plus a `raw` sheet of the filtered per-run rows

    python scripts/build_ett_expert_sweep.py
"""
import csv, os, collections, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IN_CSV = os.path.join(ROOT, 'docs', 'calibration_results_tsf.csv')
OUT_CSV = os.path.join(ROOT, 'docs', 'ett_expert_sweep_mog.csv')
OUT = os.path.join(ROOT, 'docs', 'ett_expert_sweep_mog.xlsx')

ETT = ['ETTh1', 'ETTh2', 'ETTm1', 'ETTm2']
NE = ['1', '2', '3', '4', '5']
TARGET = 0.90

METHOD_ORDER = ['standard_cp', 'aci_cp', 'cpvs', 'aci_cpvs',
                'aleatoric_only', 'aci_aleatoric_only', 'cp_aleatoric_scale',
                'adaptive_window_cp', 'moecp', 'aci_moecp',
                'aci_aleatoric_scale', 'aci_aleatoric_scale_g001']
# Paper-facing names: cp_aleatoric_scale is CP-MoG and aleatoric_only is CPVS-aleatoric,
# matching the column names in docs/table_seed*.csv.
LABEL = {'standard_cp': 'CP', 'cpvs': 'CPVS', 'aleatoric_only': 'CPVS-aleatoric',
         'cp_aleatoric_scale': 'CP-MoG',
         'adaptive_window_cp': 'Adaptive window CP', 'moecp': 'MoECP',
         'aci_aleatoric_scale': 'CP-MoG + ACI (g=0.01)',
         'aci_aleatoric_scale_g001': 'CP-MoG + ACI (g=0.001)',
         'aci_cp': 'CP + ACI (g=0.001)', 'aci_cpvs': 'CPVS + ACI (g=0.001)',
         'aci_aleatoric_only': 'CPVS-aleatoric + ACI (g=0.001)',
         'aci_moecp': 'MoECP + ACI (g=0.001)'}

rows = [r for r in csv.DictReader(open(IN_CSV))
        if r['dataset'] in ETT and r['variant'] == 'MOG' and r['model'] == 'iTransformer'
        and r['model_id'] == 'test' and r['features'] == 'M' and r['num_experts'] in NE
        and r['coverage']]

for m in sorted({r['method'] for r in rows}):
    if m not in METHOD_ORDER:
        METHOD_ORDER.append(m)
        LABEL.setdefault(m, m.replace('_', ' ').title())


def cell(rs):
    """(mean coverage, mean width, n) over a set of runs; width may be nan."""
    covs = [float(r['coverage']) for r in rs]
    wids = [float(r['width']) for r in rs if r['width']]
    if not covs:
        return None
    return (sum(covs) / len(covs), sum(wids) / len(wids) if wids else None, len(covs))


def balanced_methods(pr, methods):
    """Methods covering all 4 datasets x 5 expert counts with the same seed count per cell.

    Anything else has a trend across experts that is partly a difference in which configs ran,
    so both outputs flag it rather than presenting it next to the balanced rows unmarked.
    """
    out = []
    for m in methods:
        sizes = {len({r['seed'] for r in pr if r['method'] == m and r['dataset'] == ds
                      and r['num_experts'] == ne}) for ds in ETT for ne in NE}
        if len(sizes) == 1 and 0 not in sizes:
            out.append(m)
    return out


def mse_mae(pr, ds, ne):
    """Mean MSE/MAE of the trained configs behind a (dataset, n_exp) cell.

    Keyed by setting first: the same trained model appears once per calibrator, and counting it
    once per method would weight configs by how many calibrators happened to run on them.
    """
    per_cfg = {r['setting']: (float(r['mse']), float(r['mae'])) for r in pr
               if r['dataset'] == ds and r['num_experts'] == ne and r['mse']}
    v = list(per_cfg.values())
    if not v:
        return (None, None)
    return (sum(t[0] for t in v) / len(v), sum(t[1] for t in v) / len(v))


# ------------------------------------------------------------------- tidy csv
sweep_pls = [pl for pl in sorted({r['pred_len'] for r in rows}, key=int)
             if {r['num_experts'] for r in rows if r['pred_len'] == pl} == set(NE)]

CSV_COLS = ['pred_len', 'dataset', 'method', 'method_label', 'balanced_grid', 'num_experts',
            'n_runs', 'coverage', 'coverage_gap', 'width', 'width_vs_ne1', 'mse', 'mae',
            'target_coverage', 'variant', 'model', 'features', 'seq_len', 'alpha', 'seeds']


def fmt(v, nd=4):
    return '' if v is None else f'{v:.{nd}f}'


csv_rows = []
for pl in sweep_pls:
    pr = [r for r in rows if r['pred_len'] == pl]
    methods = [m for m in METHOD_ORDER if any(r['method'] == m for r in pr)]
    full = balanced_methods(pr, methods)
    for m in methods:
        for ds in ['ALL'] + ETT:
            scope = pr if ds == 'ALL' else [r for r in pr if r['dataset'] == ds]
            base = {d: cell([r for r in scope if r['method'] == m and r['num_experts'] == '1'
                             and r['dataset'] == d]) for d in ETT}
            for ne in NE:
                sub = [r for r in scope if r['method'] == m and r['num_experts'] == ne]
                a = cell(sub)
                if not a:
                    continue
                # Raw widths live on four different target scales, so the ratio to ne=1 is
                # always taken inside a dataset and only then averaged.
                ratios = []
                for d in (ETT if ds == 'ALL' else [ds]):
                    x = cell([r for r in pr if r['method'] == m and r['num_experts'] == ne
                              and r['dataset'] == d])
                    b = base[d]
                    if x and b and x[1] and b[1]:
                        ratios.append(x[1] / b[1])
                rel = sum(ratios) / len(ratios) if len(ratios) == (4 if ds == 'ALL' else 1) else None
                mse, mae = ((None, None) if ds == 'ALL' else mse_mae(pr, ds, ne))
                if ds == 'ALL':
                    per = [mse_mae(pr, d, ne) for d in ETT]
                    per = [t for t in per if t[0] is not None]
                    if per:
                        mse = sum(t[0] for t in per) / len(per)
                        mae = sum(t[1] for t in per) / len(per)
                csv_rows.append({
                    'pred_len': pl, 'dataset': ds, 'method': m, 'method_label': LABEL[m],
                    'balanced_grid': int(m in full), 'num_experts': ne, 'n_runs': a[2],
                    'coverage': fmt(a[0]), 'coverage_gap': fmt(a[0] - TARGET),
                    'width': fmt(a[1]), 'width_vs_ne1': fmt(rel, 4),
                    'mse': fmt(mse, 6), 'mae': fmt(mae, 6),
                    'target_coverage': f'{TARGET:.2f}', 'variant': 'MOG', 'model': 'iTransformer',
                    'features': 'M', 'seq_len': '96', 'alpha': '0.1',
                    'seeds': ' '.join(sorted({r['seed'] for r in sub}))})

with open(OUT_CSV, 'w', newline='') as fh:
    w = csv.DictWriter(fh, fieldnames=CSV_COLS)
    w.writeheader()
    for r in sorted(csv_rows, key=lambda x: (int(x['pred_len']),
                                             ETT.index(x['dataset']) if x['dataset'] in ETT else -1,
                                             1 - x['balanced_grid'], METHOD_ORDER.index(x['method']),
                                             int(x['num_experts']))):
        w.writerow(r)
print(f'wrote {OUT_CSV}: {len(csv_rows)} rows')


# --------------------------------------------------------------------- styling
HDR = Font(bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='44546A')
TITLE = Font(bold=True, size=12)
SUB = Font(bold=True)
GREY = PatternFill('solid', fgColor='F2F2F2')


def write_block(ws, r0, title, header, body, fmts):
    """One titled table starting at row r0; returns the next free row."""
    ws.cell(r0, 1, title).font = TITLE
    r = r0 + 1
    for j, h in enumerate(header, 1):
        c = ws.cell(r, j, h)
        c.font, c.fill = HDR, HDR_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, line in enumerate(body):
        r += 1
        for j, v in enumerate(line, 1):
            c = ws.cell(r, j, v)
            if j > len(fmts) - 1 or fmts[j - 1] is None:
                pass
            if isinstance(v, float):
                c.number_format = fmts[j - 1] or '0.0000'
            if i % 2:
                c.fill = GREY
    return r + 3


def autosize(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


wb = Workbook()

# ------------------------------------------------------------------ README
ws = wb.active
ws.title = 'README'
info = [
    ('ETT expert sweep — MOG, num_experts 1..5', ''),
    ('', ''),
    ('generated', datetime.date.today().isoformat()),
    ('source', 'docs/calibration_results_tsf.csv (scripts/collect_calibration_results.py)'),
    ('mirrors', 'section 6 of docs/calibration_results_tsf.md'),
    ('', ''),
    ('datasets', ', '.join(ETT)),
    ('variant', 'MOG — prob_expert=1, unc_gating=0'),
    ('backbone', 'iTransformer'),
    ('features', 'M (multivariate)'),
    ('model_id', 'test'),
    ('seq_len', '96'),
    ('alpha', f'0.1 — target coverage {TARGET:.2f}'),
    ('num_experts', ', '.join(NE)),
    ('', ''),
    ('coverage', 'fraction of test timesteps inside the interval; closer to 0.90 is better'),
    ('width', 'mean interval width in target units; lower is better at equal coverage'),
    ('width vs ne=1', 'per-dataset width ratio to the 1-expert model, then averaged over the '
                      '4 datasets — raw widths live on 4 different scales, so they are not '
                      'averaged directly. <1.00 = tighter than 1 expert.'),
    ('runs', 'number of (dataset, seed) runs behind the cell'),
    ('* on a method', 'that method does NOT cover all 4 datasets x 5 expert counts with an equal '
                      'number of seeds per cell, so its trend across experts is partly a '
                      'difference in which configs ran'),
]
for i, (k, v) in enumerate(info, 1):
    ws.cell(i, 1, k).font = TITLE if i == 1 else SUB
    ws.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical='top')
autosize(ws, [18, 95])

# ------------------------------------------------------- one sheet per horizon
for pl in sweep_pls:
    pr = [r for r in rows if r['pred_len'] == pl]
    methods = [m for m in METHOD_ORDER if any(r['method'] == m for r in pr)]
    full = balanced_methods(pr, methods)
    ordered = full + [m for m in methods if m not in full]

    def name(m):
        return LABEL[m] + ('' if m in full else ' *')

    ws = wb.create_sheet(f'pl{pl}')
    hdr = ['method'] + [f'ne={n}' for n in NE]
    r = 1
    ws.cell(r, 1, f'pred_len {pl} — {len(pr)} results, seeds '
                  f'{", ".join(sorted({x["seed"] for x in pr}))}, '
                  f'{len(full)}/{len(methods)} methods on a balanced grid').font = TITLE
    r += 2

    body = [[name(m)] + [(lambda a: a[0] if a else None)(
        cell([x for x in pr if x['method'] == m and x['num_experts'] == ne])) for ne in NE]
        for m in ordered]
    r = write_block(ws, r, f'Mean coverage over the 4 ETT datasets (target {TARGET:.2f})',
                    hdr, body, [None] + ['0.0000'] * 5)

    body = [[name(m)] + [(lambda a: a[1] if a else None)(
        cell([x for x in pr if x['method'] == m and x['num_experts'] == ne])) for ne in NE]
        for m in ordered]
    r = write_block(ws, r, 'Mean width over the 4 ETT datasets',
                    hdr, body, [None] + ['0.0000'] * 5)

    body = []
    for m in ordered:
        line = [name(m)]
        for ne in NE:
            ratios = []
            for ds in ETT:
                a = cell([x for x in pr if x['method'] == m and x['num_experts'] == ne and x['dataset'] == ds])
                b = cell([x for x in pr if x['method'] == m and x['num_experts'] == '1' and x['dataset'] == ds])
                if a and b and a[1] and b[1]:
                    ratios.append(a[1] / b[1])
            line.append(sum(ratios) / len(ratios) if len(ratios) == len(ETT) else None)
        body.append(line)
    r = write_block(ws, r, 'Width relative to ne=1 (per-dataset ratio, then averaged)',
                    hdr, body, [None] + ['0.000'] * 5)

    body = [[name(m)] + [(lambda a: a[2] if a else 0)(
        cell([x for x in pr if x['method'] == m and x['num_experts'] == ne])) for ne in NE]
        for m in ordered]
    r = write_block(ws, r, 'Runs behind each cell (4 datasets x seeds)',
                    hdr, body, [None] + ['0'] * 5)

    for metric, idx in (('MSE', 0), ('MAE', 1)):
        body = [[ds] + [mse_mae(pr, ds, ne)[idx] for ne in NE] for ds in ETT]
        r = write_block(ws, r, f'Point accuracy of the underlying models — {metric}',
                        ['dataset'] + [f'ne={n}' for n in NE], body, [None] + ['0.0000'] * 5)

    autosize(ws, [30] + [12] * 5)
    ws.freeze_panes = 'B1'

    # ------------------------------------------------- per-dataset detail sheet
    ws = wb.create_sheet(f'pl{pl}_per_dataset')
    hdr = ['dataset', 'method', 'metric'] + [f'ne={n}' for n in NE]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(1, j, h)
        c.font, c.fill = HDR, HDR_FILL
        c.alignment = Alignment(horizontal='center')
    r = 1
    for ds in ETT:
        for m in ordered:
            vals = {ne: cell([x for x in pr if x['method'] == m and x['num_experts'] == ne
                              and x['dataset'] == ds]) for ne in NE}
            if not any(vals.values()):
                continue
            for metric, idx, fmt in (('coverage', 0, '0.0000'), ('width', 1, '0.0000'),
                                     ('runs', 2, '0')):
                r += 1
                ws.cell(r, 1, ds)
                ws.cell(r, 2, name(m))
                ws.cell(r, 3, metric)
                for j, ne in enumerate(NE, 4):
                    a = vals[ne]
                    c = ws.cell(r, j, a[idx] if a else None)
                    c.number_format = fmt
                if metric == 'coverage':
                    for j in range(1, len(hdr) + 1):
                        ws.cell(r, j).fill = GREY
    autosize(ws, [10, 30, 10] + [12] * 5)
    ws.freeze_panes = 'D2'

# ------------------------------------------------------------------- raw sheet
ws = wb.create_sheet('raw')
raw_cols = ['dataset', 'num_experts', 'pred_len', 'seed', 'method', 'coverage', 'width',
            'median_width', 'mse', 'mae', 'variant', 'model', 'alpha', 'seq_len', 'setting']
for j, h in enumerate(raw_cols, 1):
    c = ws.cell(1, j, h)
    c.font, c.fill = HDR, HDR_FILL
for i, x in enumerate(sorted(rows, key=lambda x: (x['dataset'], int(x['pred_len']),
                                                  int(x['num_experts']), x['method'], x['seed'])), 2):
    for j, k in enumerate(raw_cols, 1):
        v = x.get(k, '')
        if k in ('coverage', 'width', 'median_width', 'mse', 'mae', 'alpha'):
            v = float(v) if v else None
        elif k in ('num_experts', 'pred_len', 'seq_len'):
            v = int(v) if v else None
        ws.cell(i, j, v)
autosize(ws, [10, 12, 10, 8, 24, 11, 11, 13, 10, 10, 9, 14, 8, 9, 60])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(raw_cols))}{len(rows) + 1}'

wb.save(OUT)
print(f'wrote {OUT}: {len(wb.sheetnames)} sheets ({", ".join(wb.sheetnames)}), '
      f'{len(rows)} raw runs')
